import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from config import file_path
from config import  output_file
from openpyxl import Workbook


# Load the CSV file
data = pd.read_csv(file_path)

# Ensure the "Time" column is properly parsed as datetime
data['Time'] = pd.to_datetime(data['Time'], errors='coerce')

# Extract "Month-Year" from the "Time" column
data['Month-Year'] = data['Time'].dt.strftime('%m.%Y')

# Group data by "Month-Year" and calculate Profit and Loss metrics
monthly_summary = data.groupby('Month-Year').agg(
    Total_Profit=('Result', lambda x: x[x > 0].sum()),
    Total_Loss=('Result', lambda x: x[x < 0].sum()),
    Positive_transactions=('Result', lambda x: (x > 0).sum()),
    Negative_transactions=('Result', lambda x: (x < 0).sum()),
    Total_Currency_Conversion_Fees=('Currency conversion fee', 'sum')
).reset_index()

# Ensure currency conversion fees are always positive by taking the absolute value, if needed
monthly_summary['Total_Currency_Conversion_Fees'] = monthly_summary['Total_Currency_Conversion_Fees'].abs()

# Calculate Profit and Loss: Total Profit - Total Loss - Total Currency Conversion Fees
monthly_summary['Profit_and_Loss'] = monthly_summary['Total_Profit'] + monthly_summary['Total_Loss'] - monthly_summary['Total_Currency_Conversion_Fees']

# Reorder columns to place "Profit_and_Loss" second
columns_order = ['Month-Year', 'Profit_and_Loss', 'Total_Profit', 'Total_Loss', 'Positive_transactions', 'Negative_transactions', 'Total_Currency_Conversion_Fees']
monthly_summary = monthly_summary[columns_order]

# Rename columns by replacing underscores with spaces
monthly_summary.columns = [col.replace('_', ' ') for col in monthly_summary.columns]


# Define the list of specific names
keywords = [
    "Vanguard S&P 500 (Dist)",
    "Vanguard FTSE All-World (Acc)",
    "iShares MSCI India (Acc)",
    "iShares S&P 500 Consumer Staples Sector"
]

# Filter data matching these keywords in the "Name" column
filtered_data = data[data['Name'].isin(keywords)]

# Calculate the total yearly profit for the filtered data
yearly_profit_sum = filtered_data['Result'].sum()

# Group filtered data by "Name" to calculate total profits for each keyword
namewise_summary = filtered_data.groupby('Name').agg(
    Total_Profit=('Result', 'sum')
).reset_index()



# Save the monthly summary to a new Excel file
# Add a final row for total profit across all keywords
total_row = pd.DataFrame({'Name': ['Total'], 'Total_Profit': [yearly_profit_sum]})
namewise_summary = pd.concat([namewise_summary, total_row], ignore_index=True)



# Check if the output file exists
if not os.path.exists(output_file):
    wb = Workbook()  # Create a new workbook
    # Remove the default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb['Sheet']
    wb.save(output_file)  # Save it so that it exists for subsequent operations
else:
    wb = openpyxl.load_workbook(output_file)


# Save the monthly summary to a new sheet
monthly_summary_sheet_name = "Monthly Summary"
if monthly_summary_sheet_name in wb.sheetnames:
    del wb[monthly_summary_sheet_name]  # Remove the sheet if it already exists
ws_monthly = wb.create_sheet(title=monthly_summary_sheet_name)
# Write the monthly summary to the new sheet
for row in dataframe_to_rows(monthly_summary, index=False, header=True):
    ws_monthly.append(row)


if not filtered_data.empty:
    # Create a new sheet for name-based yearly profits
    summary_sheet_name = "UCITS Yearly Profits by Name"
    if summary_sheet_name in wb.sheetnames:
        del wb[summary_sheet_name]  # Remove the sheet if it already exists
    ws = wb.create_sheet(title=summary_sheet_name)

    # Write the summary to the new sheet
    for row in dataframe_to_rows(namewise_summary, index=False, header=True):
        ws.append(row)



# Autofit columns dynamically based on the maximum length of content in each column
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # Adding extra space for padding
    ws.column_dimensions[column].width = adjusted_width

# Center-align all cells in the worksheet
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

# Save the file again after autofitting the columns and aligning the cells
# Save the updated workbook
wb.save(output_file)

# Open the Excel file using the default viewer
os.startfile(output_file)

print(f"Monthly Profit and Loss statement saved to {output_file}")
print(f"Yearly profits for specific names saved to sheet: {summary_sheet_name} in {output_file}")